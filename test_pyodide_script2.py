import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart.series import DataPoint
import pandas as pd
from datetime import datetime

# Create workbook
wb = Workbook()

# Define styles
header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=14, color="1E3A5F")
subtitle_font = Font(bold=True, size=11, color="2E5077")
money_font = Font(bold=True, color="1E3A5F")
positive_font = Font(color="2E7D32")
negative_font = Font(color="C62828")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
light_blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
light_green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
light_red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

# ==========================================
# SHEET 1: DASHBOARD
# ==========================================
ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

# Title
ws_dashboard.merge_cells('A1:G1')
ws_dashboard['A1'] = "PERSONAL FINANCE DASHBOARD"
ws_dashboard['A1'].font = Font(bold=True, size=20, color="1E3A5F")
ws_dashboard['A1'].alignment = Alignment(horizontal='center')

ws_dashboard['A2'] = f"Last Updated: {datetime.now().strftime('%B %d, %Y')}"
ws_dashboard['A2'].font = Font(italic=True, size=10)
ws_dashboard['A2'].alignment = Alignment(horizontal='center')

# Monthly Summary
ws_dashboard['A4'] = "MONTHLY SUMMARY"
ws_dashboard['A4'].font = title_font

summary_items = [
    ('Total Income', '=Income!B20', light_green_fill),
    ('Total Expenses', '=Expenses!B18', light_red_fill),
    ('Net Savings', '=B6-B7', light_blue_fill),
    ('Savings Rate', '=IF(B6>0,(B8/B6)*100,0)', light_yellow_fill),
]

row = 5
for label, formula, fill in summary_items:
    ws_dashboard[f'A{row}'] = label
    ws_dashboard[f'A{row}'].font = Font(bold=True)
    ws_dashboard[f'A{row}'].fill = fill
    ws_dashboard[f'B{row}'] = formula
    ws_dashboard[f'B{row}'].number_format = '"$"#,##0.00'
    ws_dashboard[f'B{row}'].fill = fill
    ws_dashboard[f'B{row}'].font = money_font
    row += 1

ws_dashboard['B9'].number_format = '0.00"%"'

# Quick Stats
ws_dashboard['A11'] = "FINANCIAL HEALTH"
ws_dashboard['A11'].font = title_font

health_items = [
    ('Active Loans', '=Loans!B14'),
    ('Total Loan Balance', '=Loans!B15'),
    ('Savings Goals Progress', '=Savings!B12'),
    ('Budget Status', '=IF(Dashboard!B8>0,IF(Dashboard!B8/Dashboard!B6>0.2,"On Track","Needs Attention"),"No Data")'),
]

row = 12
for label, formula in health_items:
    ws_dashboard[f'A{row}'] = label
    ws_dashboard[f'A{row}'].font = Font(bold=True)
    ws_dashboard[f'B{row}'] = formula
    ws_dashboard[f'B{row}'].fill = light_blue_fill
    row += 1

ws_dashboard['B15'].number_format = '"$"#,##0.00'

# Column widths
ws_dashboard.column_dimensions['A'].width = 25
ws_dashboard.column_dimensions['B'].width = 18

# ==========================================
# SHEET 2: INCOME & BUDGET
# ==========================================
ws_income = wb.create_sheet("Income")

ws_income.merge_cells('A1:F1')
ws_income['A1'] = "MONTHLY INCOME & BUDGET"
ws_income['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_income['A1'].alignment = Alignment(horizontal='center')

# Income Sources
ws_income['A3'] = "INCOME SOURCES"
ws_income['A3'].font = subtitle_font

income_headers = ['Source', 'Budgeted', 'Actual', 'Variance', 'Notes']
for col, header in enumerate(income_headers, 1):
    cell = ws_income.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

income_sources = [
    ('Salary', 5000, '', '=C5-B5', 'Primary income'),
    ('Freelance', 500, '', '=C6-B6', ''),
    ('Investments', 200, '', '=C7-B7', ''),
    ('Side Business', 300, '', '=C8-B8', ''),
    ('Other', 100, '', '=C9-B9', ''),
]

for row_idx, (source, budget, actual, variance, notes) in enumerate(income_sources, 5):
    ws_income.cell(row=row_idx, column=1, value=source).border = thin_border
    ws_income.cell(row=row_idx, column=2, value=budget).number_format = '"$"#,##0.00'
    ws_income.cell(row=row_idx, column=2).border = thin_border
    ws_income.cell(row=row_idx, column=3, value=actual).number_format = '"$"#,##0.00'
    ws_income.cell(row=row_idx, column=3).border = thin_border
    ws_income.cell(row=row_idx, column=4, value=variance).number_format = '"$"#,##0.00'
    ws_income.cell(row=row_idx, column=4).border = thin_border
    ws_income.cell(row=row_idx, column=5, value=notes).border = thin_border

# Totals
ws_income['A10'] = "Total Income"
ws_income['A10'].font = Font(bold=True)
ws_income['B10'] = '=SUM(B5:B9)'
ws_income['B10'].number_format = '"$"#,##0.00'
ws_income['B10'].font = money_font
ws_income['C10'] = '=SUM(C5:C9)'
ws_income['C10'].number_format = '"$"#,##0.00'
ws_income['C10'].font = money_font
ws_income['D10'] = '=C10-B10'
ws_income['D10'].number_format = '"$"#,##0.00'

# Budget Categories
ws_income['A12'] = "BUDGET CATEGORIES"
ws_income['A12'].font = subtitle_font

budget_headers = ['Category', 'Budgeted', 'Spent', 'Remaining', 'Status']
for col, header in enumerate(budget_headers, 1):
    cell = ws_income.cell(row=13, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

budget_categories = [
    ('Housing', 1500),
    ('Utilities', 200),
    ('Groceries', 600),
    ('Transportation', 400),
    ('Insurance', 300),
    ('Entertainment', 200),
    ('Dining Out', 250),
    ('Healthcare', 150),
    ('Personal Care', 100),
    ('Education', 100),
    ('Savings', 500),
    ('Miscellaneous', 200),
]

for row_idx, (cat, budget) in enumerate(budget_categories, 14):
    ws_income.cell(row=row_idx, column=1, value=cat).border = thin_border
    ws_income.cell(row=row_idx, column=2, value=budget).number_format = '"$"#,##0.00'
    ws_income.cell(row=row_idx, column=2).border = thin_border
    ws_income.cell(row=row_idx, column=3).border = thin_border
    ws_income.cell(row=row_idx, column=4, value=f'=B{row_idx}-C{row_idx}')
    ws_income.cell(row=row_idx, column=4).number_format = '"$"#,##0.00'
    ws_income.cell(row=row_idx, column=4).border = thin_border
    ws_income.cell(row=row_idx, column=5, value=f'=IF(C{row_idx}>B{row_idx},"Over Budget",IF(C{row_idx}=B{row_idx},"On Target","Under Budget"))')
    ws_income.cell(row=row_idx, column=5).border = thin_border

# Budget Totals
ws_income['A26'] = "Total Budget"
ws_income['A26'].font = Font(bold=True)
ws_income['B26'] = '=SUM(B14:B25)'
ws_income['B26'].number_format = '"$"#,##0.00'
ws_income['B26'].font = money_font
ws_income['C26'] = '=SUM(C14:C25)'
ws_income['C26'].number_format = '"$"#,##0.00'
ws_income['C26'].font = money_font
ws_income['D26'] = '=B26-C26'
ws_income['D26'].number_format = '"$"#,##0.00'

# Summary row
ws_income['A28'] = "MONTHLY SUMMARY"
ws_income['A28'].font = subtitle_font
ws_income['A29'] = "Total Income"
ws_income['B29'] = '=B10'
ws_income['B29'].number_format = '"$"#,##0.00'
ws_income['A30'] = "Total Budget"
ws_income['B30'] = '=B26'
ws_income['B30'].number_format = '"$"#,##0.00'
ws_income['A31'] = "Remaining"
ws_income['B31'] = '=B29-B30'
ws_income['B31'].number_format = '"$"#,##0.00'
ws_income['B31'].font = money_font

# Column widths
for col in range(1, 6):
    ws_income.column_dimensions[get_column_letter(col)].width = 18

# ==========================================
# SHEET 3: EXPENSES TRACKER
# ==========================================
ws_expenses = wb.create_sheet("Expenses")

ws_expenses.merge_cells('A1:H1')
ws_expenses['A1'] = "EXPENSES TRACKER"
ws_expenses['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_expenses['A1'].alignment = Alignment(horizontal='center')

expense_headers = ['Date', 'Category', 'Description', 'Amount', 'Payment Method', 'Receipt', 'Notes', 'Month']
for col, header in enumerate(expense_headers, 1):
    cell = ws_expenses.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Sample expense data
sample_expenses = [
    ('2026-06-01', 'Housing', 'Monthly Rent', 1500, 'Bank Transfer', 'Yes', '', 'June'),
    ('2026-06-02', 'Groceries', 'Weekly Shopping', 145.50, 'Credit Card', 'Yes', '', 'June'),
    ('2026-06-03', 'Utilities', 'Electric Bill', 85.20, 'Auto Pay', 'Yes', '', 'June'),
    ('2026-06-05', 'Transportation', 'Gas', 55.00, 'Debit Card', 'No', '', 'June'),
    ('2026-06-07', 'Dining Out', 'Restaurant', 42.50, 'Credit Card', 'No', '', 'June'),
    ('2026-06-10', 'Healthcare', 'Pharmacy', 28.90, 'Debit Card', 'Yes', '', 'June'),
    ('2026-06-12', 'Entertainment', 'Movie Tickets', 32.00, 'Credit Card', 'No', '', 'June'),
    ('2026-06-15', 'Groceries', 'Weekly Shopping', 132.75, 'Credit Card', 'Yes', '', 'June'),
    ('2026-06-18', 'Personal Care', 'Haircut', 45.00, 'Cash', 'No', '', 'June'),
    ('2026-06-20', 'Utilities', 'Internet', 65.00, 'Auto Pay', 'Yes', '', 'June'),
    ('2026-06-22', 'Dining Out', 'Coffee Shop', 18.50, 'Credit Card', 'No', '', 'June'),
    ('2026-06-25', 'Transportation', 'Parking', 25.00, 'Debit Card', 'No', '', 'June'),
]

for row_idx, expense in enumerate(sample_expenses, 4):
    for col_idx, value in enumerate(expense, 1):
        cell = ws_expenses.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 4:
            cell.number_format = '"$"#,##0.00'

# Totals
ws_expenses['A17'] = "TOTAL EXPENSES"
ws_expenses['A17'].font = Font(bold=True)
ws_expenses['D17'] = '=SUM(D4:D16)'
ws_expenses['D17'].number_format = '"$"#,##0.00'
ws_expenses['D17'].font = money_font

# Category Summary
ws_expenses['A19'] = "EXPENSES BY CATEGORY"
ws_expenses['A19'].font = subtitle_font

cat_summary_headers = ['Category', 'Total Spent', '% of Total', 'Budget', 'Remaining']
for col, header in enumerate(cat_summary_headers, 1):
    cell = ws_expenses.cell(row=20, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

categories = ['Housing', 'Utilities', 'Groceries', 'Transportation', 'Insurance', 
              'Entertainment', 'Dining Out', 'Healthcare', 'Personal Care', 'Education', 'Miscellaneous']

for row_idx, cat in enumerate(categories, 21):
    ws_expenses.cell(row=row_idx, column=1, value=cat).border = thin_border
    ws_expenses.cell(row=row_idx, column=2, value=f'=SUMIF(B:B,A{row_idx},D:D)')
    ws_expenses.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
    ws_expenses.cell(row=row_idx, column=2).border = thin_border
    ws_expenses.cell(row=row_idx, column=3, value=f'=IF(D17>0,B{row_idx}/D17*100,0)')
    ws_expenses.cell(row=row_idx, column=3).number_format = '0.00"%"'
    ws_expenses.cell(row=row_idx, column=3).border = thin_border
    ws_expenses.cell(row=row_idx, column=4, value=f'=VLOOKUP(A{row_idx},Income!A14:B25,2,FALSE)')
    ws_expenses.cell(row=row_idx, column=4).number_format = '"$"#,##0.00'
    ws_expenses.cell(row=row_idx, column=4).border = thin_border
    ws_expenses.cell(row=row_idx, column=5, value=f'=D{row_idx}-B{row_idx}')
    ws_expenses.cell(row=row_idx, column=5).number_format = '"$"#,##0.00'
    ws_expenses.cell(row=row_idx, column=5).border = thin_border

# Column widths
ws_expenses.column_dimensions['A'].width = 12
ws_expenses.column_dimensions['B'].width = 18
ws_expenses.column_dimensions['C'].width = 25
ws_expenses.column_dimensions['D'].width = 12
ws_expenses.column_dimensions['E'].width = 15
ws_expenses.column_dimensions['F'].width = 10
ws_expenses.column_dimensions['G'].width = 15
ws_expenses.column_dimensions['H'].width = 10

# ==========================================
# SHEET 4: LOANS MANAGER
# ==========================================
ws_loans = wb.create_sheet("Loans")

ws_loans.merge_cells('A1:J1')
ws_loans['A1'] = "LOANS MANAGER"
ws_loans['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_loans['A1'].alignment = Alignment(horizontal='center')

# Active Loans
ws_loans['A3'] = "ACTIVE LOANS"
ws_loans['A3'].font = subtitle_font

loan_headers = ['Loan Name', 'Type', 'Principal', 'Interest Rate', 'Term', 'Monthly Payment', 
                'Start Date', 'Remaining Payments', 'Total Paid', 'Remaining Balance']
for col, header in enumerate(loan_headers, 1):
    cell = ws_loans.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Sample loan data
loans_data = [
    ('Car Loan', 'Auto', 25000, 5.5, 60, 477.00, '2024-01-15', 42, 4770, 20025),
    ('Student Loan', 'Education', 35000, 4.2, 120, 357.00, '2023-06-01', 96, 8568, 29856),
    ('Credit Card 1', 'Credit', 3000, 19.99, 24, 150.00, '2025-03-01', 21, 1350, 1950),
]

for row_idx, loan in enumerate(loans_data, 5):
    for col_idx, value in enumerate(loan, 1):
        cell = ws_loans.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [3, 6, 9, 10]:
            cell.number_format = '"$"#,##0.00'
        if col_idx == 4:
            cell.number_format = '0.00"%"'

# Loan Totals
ws_loans['A9'] = "TOTALS"
ws_loans['A9'].font = Font(bold=True)
ws_loans['C9'] = '=SUM(C5:C7)'
ws_loans['C9'].number_format = '"$"#,##0.00'
ws_loans['C9'].font = money_font
ws_loans['F9'] = '=SUM(F5:F7)'
ws_loans['F9'].number_format = '"$"#,##0.00'
ws_loans['F9'].font = money_font
ws_loans['I9'] = '=SUM(I5:I7)'
ws_loans['I9'].number_format = '"$"#,##0.00'
ws_loans['J9'] = '=SUM(J5:J7)'
ws_loans['J9'].number_format = '"$"#,##0.00'
ws_loans['J9'].font = money_font

# Active Loans Count
ws_loans['A11'] = "Active Loans"
ws_loans['B11'] = '=COUNTA(A5:A7)'
ws_loans['A12'] = "Total Outstanding Balance"
ws_loans['B12'] = '=J9'
ws_loans['B12'].number_format = '"$"#,##0.00'
ws_loans['A13'] = "Total Monthly Payments"
ws_loans['B13'] = '=F9'
ws_loans['B13'].number_format = '"$"#,##0.00'
ws_loans['A14'] = '=B11'
ws_loans['B15'] = '=B12'
ws_loans['B15'].number_format = '"$"#,##0.00'

# Loan Payoff Timeline
ws_loans['A17'] = "LOAN PAYOFF TIMELINE"
ws_loans['A17'].font = subtitle_font

timeline_headers = ['Month', 'Payment', 'Principal', 'Interest', 'Balance']
for col, header in enumerate(timeline_headers, 1):
    cell = ws_loans.cell(row=18, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Amortization for first 12 months of Car Loan
for row_idx in range(19, 31):
    month = row_idx - 18
    ws_loans.cell(row=row_idx, column=1, value=f'Month {month}').border = thin_border
    ws_loans.cell(row=row_idx, column=2, value=477).number_format = '"$"#,##0.00'
    ws_loans.cell(row=row_idx, column=2).border = thin_border
    ws_loans.cell(row=row_idx, column=3, value=f'=B{row_idx}*(1-0.00458)').number_format = '"$"#,##0.00'
    ws_loans.cell(row=row_idx, column=3).border = thin_border
    ws_loans.cell(row=row_idx, column=4, value=f'=B{row_idx}-C{row_idx}').number_format = '"$"#,##0.00'
    ws_loans.cell(row=row_idx, column=4).border = thin_border
    ws_loans.cell(row=row_idx, column=5, value=f'=IF({row_idx}=19,20025-B{row_idx},E{row_idx-1}-B{row_idx})')
    ws_loans.cell(row=row_idx, column=5).number_format = '"$"#,##0.00'
    ws_loans.cell(row=row_idx, column=5).border = thin_border

# Column widths
for col in range(1, 11):
    ws_loans.column_dimensions[get_column_letter(col)].width = 15

# ==========================================
# SHEET 5: SAVINGS TRACKER
# ==========================================
ws_savings = wb.create_sheet("Savings")

ws_savings.merge_cells('A1:F1')
ws_savings['A1'] = "SAVINGS TRACKER"
ws_savings['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_savings['A1'].alignment = Alignment(horizontal='center')

# Savings Goals
ws_savings['A3'] = "SAVINGS GOALS"
ws_savings['A3'].font = subtitle_font

savings_headers = ['Goal Name', 'Target Amount', 'Current Amount', 'Monthly Contribution', 'Target Date', 'Progress %']
for col, header in enumerate(savings_headers, 1):
    cell = ws_savings.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

goals_data = [
    ('Emergency Fund', 15000, 8500, 500, '2026-12-31'),
    ('Vacation', 5000, 1200, 300, '2026-08-01'),
    ('New Car Down Payment', 10000, 3500, 400, '2027-06-01'),
    ('Home Down Payment', 50000, 15000, 800, '2029-01-01'),
    ('Retirement', 500000, 45000, 600, '2045-01-01'),
]

for row_idx, goal in enumerate(goals_data, 5):
    for col_idx, value in enumerate(goal, 1):
        cell = ws_savings.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [2, 3, 4]:
            cell.number_format = '"$"#,##0.00'
    ws_savings.cell(row=row_idx, column=6, value=f'=C{row_idx}/B{row_idx}*100')
    ws_savings.cell(row=row_idx, column=6).number_format = '0.00"%"'
    ws_savings.cell(row=row_idx, column=6).border = thin_border

# Totals
ws_savings['A11'] = "TOTALS"
ws_savings['A11'].font = Font(bold=True)
ws_savings['B11'] = '=SUM(B5:B9)'
ws_savings['B11'].number_format = '"$"#,##0.00'
ws_savings['B11'].font = money_font
ws_savings['C11'] = '=SUM(C5:C9)'
ws_savings['C11'].number_format = '"$"#,##0.00'
ws_savings['C11'].font = money_font
ws_savings['D11'] = '=SUM(D5:D9)'
ws_savings['D11'].number_format = '"$"#,##0.00'
ws_savings['D11'].font = money_font

ws_savings['A12'] = "Overall Progress"
ws_savings['B12'] = '=C11/B11*100'
ws_savings['B12'].number_format = '0.00"%"'
ws_savings['B12'].font = money_font

# Transaction History
ws_savings['A14'] = "SAVINGS TRANSACTIONS"
ws_savings['A14'].font = subtitle_font

trans_headers = ['Date', 'Goal', 'Type', 'Amount', 'Notes']
for col, header in enumerate(trans_headers, 1):
    cell = ws_savings.cell(row=15, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

transactions = [
    ('2026-06-01', 'Emergency Fund', 'Deposit', 500, 'Monthly contribution'),
    ('2026-06-01', 'Vacation', 'Deposit', 300, 'Monthly contribution'),
    ('2026-06-01', 'New Car Down Payment', 'Deposit', 400, 'Monthly contribution'),
    ('2026-06-05', 'Emergency Fund', 'Interest', 25.50, 'Interest earned'),
    ('2026-06-15', 'Home Down Payment', 'Deposit', 800, 'Monthly contribution'),
]

for row_idx, trans in enumerate(transactions, 16):
    for col_idx, value in enumerate(trans, 1):
        cell = ws_savings.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 4:
            cell.number_format = '"$"#,##0.00'

# Column widths
ws_savings.column_dimensions['A'].width = 20
ws_savings.column_dimensions['B'].width = 18
ws_savings.column_dimensions['C'].width = 18
ws_savings.column_dimensions['D'].width = 18
ws_savings.column_dimensions['E'].width = 15
ws_savings.column_dimensions['F'].width = 12

# ==========================================
# SHEET 6: CHARTS & VISUALIZATION
# ==========================================
ws_charts = wb.create_sheet("Charts")

ws_charts.merge_cells('A1:P1')
ws_charts['A1'] = "FINANCIAL VISUALIZATIONS"
ws_charts['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_charts['A1'].alignment = Alignment(horizontal='center')

# Expense Pie Chart Data
ws_charts['A3'] = "Expense Breakdown by Category"
ws_charts['A3'].font = subtitle_font

ws_charts['A4'] = "Category"
ws_charts['B4'] = "Amount"
ws_charts['A4'].font = header_font
ws_charts['A4'].fill = header_fill
ws_charts['B4'].font = header_font
ws_charts['B4'].fill = header_fill

expense_chart_data = [
    ('Housing', 1500),
    ('Groceries', 278),
    ('Utilities', 150),
    ('Transportation', 80),
    ('Dining Out', 61),
    ('Entertainment', 32),
    ('Healthcare', 29),
    ('Personal Care', 45),
]

for row_idx, (cat, amt) in enumerate(expense_chart_data, 5):
    ws_charts.cell(row=row_idx, column=1, value=cat)
    ws_charts.cell(row=row_idx, column=2, value=amt)
    ws_charts.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'

# Create Pie Chart for Expenses
pie = PieChart()
labels = Reference(ws_charts, min_col=1, min_row=5, max_row=12)
data = Reference(ws_charts, min_col=2, min_row=4, max_row=12)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Monthly Expenses by Category"
pie.width = 15
pie.height = 12

ws_charts.add_chart(pie, "D3")

# Monthly Income vs Expenses Bar Chart
ws_charts['A18'] = "Budget vs Actual"
ws_charts['A18'].font = subtitle_font

ws_charts['A19'] = "Category"
ws_charts['B19'] = "Budget"
ws_charts['C19'] = "Actual"
ws_charts['A19'].font = header_font
ws_charts['A19'].fill = header_fill
ws_charts['B19'].font = header_font
ws_charts['B19'].fill = header_fill
ws_charts['C19'].font = header_font
ws_charts['C19'].fill = header_fill

budget_chart_data = [
    ('Housing', 1500, 1500),
    ('Groceries', 600, 278),
    ('Utilities', 200, 150),
    ('Transportation', 400, 80),
    ('Entertainment', 200, 32),
    ('Dining Out', 250, 61),
    ('Healthcare', 150, 29),
    ('Savings', 500, 2025),
]

for row_idx, (cat, budget, actual) in enumerate(budget_chart_data, 20):
    ws_charts.cell(row=row_idx, column=1, value=cat)
    ws_charts.cell(row=row_idx, column=2, value=budget)
    ws_charts.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
    ws_charts.cell(row=row_idx, column=3, value=actual)
    ws_charts.cell(row=row_idx, column=3).number_format = '"$"#,##0.00'

# Create Bar Chart
bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.title = "Budget vs Actual Spending"
bar.y_axis.title = "Amount ($)"
bar.x_axis.title = "Category"

cats = Reference(ws_charts, min_col=1, min_row=20, max_row=27)
budget_data = Reference(ws_charts, min_col=2, min_row=19, max_row=27)
actual_data = Reference(ws_charts, min_col=3, min_row=19, max_row=27)
bar.add_data(budget_data, titles_from_data=True)
bar.add_data(actual_data, titles_from_data=True)
bar.set_categories(cats)
bar.width = 18
bar.height = 10

ws_charts.add_chart(bar, "E18")

# Savings Progress
ws_charts['A33'] = "Savings Goals Progress"
ws_charts['A33'].font = subtitle_font

ws_charts['A34'] = "Goal"
ws_charts['B34'] = "Target"
ws_charts['C34'] = "Current"
ws_chcharts = ws_charts

for col in range(1, 4):
    ws_charts.cell(row=34, column=col).font = header_font
    ws_charts.cell(row=34, column=col).fill = header_fill

savings_chart_data = [
    ('Emergency Fund', 15000, 8500),
    ('Vacation', 5000, 1200),
    ('Car Down Payment', 10000, 3500),
    ('Home Down Payment', 50000, 15000),
]

for row_idx, (goal, target, current) in enumerate(savings_chart_data, 35):
    ws_charts.cell(row=row_idx, column=1, value=goal)
    ws_charts.cell(row=row_idx, column=2, value=target)
    ws_charts.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
    ws_charts.cell(row=row_idx, column=3, value=current)
    ws_charts.cell(row=row_idx, column=3).number_format = '"$"#,##0.00'

# Create horizontal bar chart for savings
savings_bar = BarChart()
savings_bar.type = "bar"
savings_bar.grouping = "stacked"
savings_bar.title = "Savings Goals Progress"
savings_bar.y_axis.title = "Goals"
savings_bar.x_axis.title = "Amount ($)"

savings_cats = Reference(ws_charts, min_col=1, min_row=35, max_row=38)
savings_target = Reference(ws_charts, min_col=2, min_row=34, max_row=38)
savings_current = Reference(ws_charts, min_col=3, min_row=34, max_row=38)
savings_bar.add_data(savings_target, titles_from_data=True)
savings_bar.add_data(savings_current, titles_from_data=True)
savings_bar.set_categories(savings_cats)
savings_bar.width = 16
savings_bar.height = 10

ws_charts.add_chart(savings_bar, "D33")

# Column widths
for col in range(1, 4):
    ws_charts.column_dimensions[get_column_letter(col)].width = 18

# ==========================================
# SHEET 7: NET WORTH TRACKER
# ==========================================
ws_networth = wb.create_sheet("Net Worth")

ws_networth.merge_cells('A1:D1')
ws_networth['A1'] = "NET WORTH TRACKER"
ws_networth['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_networth['A1'].alignment = Alignment(horizontal='center')

# Assets
ws_networth['A3'] = "ASSETS"
ws_networth['A3'].font = subtitle_font

asset_headers = ['Asset', 'Value', 'Type', 'Notes']
for col, header in enumerate(asset_headers, 1):
    cell = ws_networth.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

assets = [
    ('Checking Account', 5000, 'Cash', ''),
    ('Savings Account', 28500, 'Cash', 'Total savings'),
    ('Emergency Fund', 8500, 'Cash', ''),
    ('401(k)', 45000, 'Investment', ''),
    ('Brokerage Account', 12000, 'Investment', ''),
    ('Car (Blue Book Value)', 18000, 'Property', ''),
    ('Home', 350000, 'Property', 'Estimated value'),
]

for row_idx, asset in enumerate(assets, 5):
    for col_idx, value in enumerate(asset, 1):
        cell = ws_networth.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 2:
            cell.number_format = '"$"#,##0.00'

# Liabilities
ws_networth['A13'] = "LIABILITIES"
ws_networth['A13'].font = subtitle_font

liability_headers = ['Liability', 'Amount', 'Type', 'Notes']
for col, header in enumerate(liability_headers, 1):
    cell = ws_networth.cell(row=14, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

liabilities = [
    ('Car Loan', 20025, 'Auto Loan', ''),
    ('Student Loan', 29856, 'Education', ''),
    ('Credit Card Balance', 1950, 'Credit', ''),
    ('Mortgage', 280000, 'Home', ''),
]

for row_idx, liab in enumerate(liabilities, 15):
    for col_idx, value in enumerate(liab, 1):
        cell = ws_networth.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 2:
            cell.number_format = '"$"#,##0.00'

# Summary
ws_networth['A20'] = "NET WORTH SUMMARY"
ws_networth['A20'].font = subtitle_font
ws_networth['A21'] = "Total Assets"
ws_networth['B21'] = '=SUM(B5:B11)'
ws_networth['B21'].number_format = '"$"#,##0.00'
ws_networth['B21'].font = money_font
ws_networth['B21'].fill = light_green_fill

ws_networth['A22'] = "Total Liabilities"
ws_networth['B22'] = '=SUM(B15:B18)'
ws_networth['B22'].number_format = '"$"#,##0.00'
ws_networth['B22'].font = Font(color="C62828", bold=True)
ws_networth['B22'].fill = light_red_fill

ws_networth['A23'] = "NET WORTH"
ws_networth['A23'].font = Font(bold=True, size=14)
ws_networth['B23'] = '=B21-B22'
ws_networth['B23'].number_format = '"$"#,##0.00'
ws_networth['B23'].font = Font(bold=True, size=14, color="1E3A5F")
ws_networth['B23'].fill = light_blue_fill

# Column widths
ws_networth.column_dimensions['A'].width = 22
ws_networth.column_dimensions['B'].width = 15
ws_networth.column_dimensions['C'].width = 15
ws_networth.column_dimensions['D'].width = 20

# ==========================================
# SHEET 8: MONTHLY OVERVIEW (BUDGET CALENDAR)
# ==========================================
ws_calendar = wb.create_sheet("Monthly Budget")

ws_calendar.merge_cells('A1:G1')
ws_calendar['A1'] = "MONTHLY BUDGET CALENDAR - JUNE 2026"
ws_calendar['A1'].font = Font(bold=True, size=18, color="1E3A5F")
ws_calendar['A1'].alignment = Alignment(horizontal='center')

# Weekly Budget Breakdown
ws_calendar['A3'] = "WEEKLY BUDGET"
ws_calendar['A3'].font = subtitle_font

week_headers = ['Week', 'Income', 'Fixed Expenses', 'Variable Expenses', 'Savings', 'Net']
for col, header in enumerate(week_headers, 1):
    cell = ws_calendar.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

weeks = [
    ('Week 1', 1500, 1850, 200, 500, '=B5-C5-D5+E5'),
    ('Week 2', 1500, 0, 350, 500, '=B6-C6-D6+E6'),
    ('Week 3', 1500, 0, 280, 500, '=B7-C7-D7+E7'),
    ('Week 4', 1500, 0, 180, 500, '=B8-C8-D8+E8'),
]

for row_idx, week in enumerate(weeks, 5):
    for col_idx, value in enumerate(week, 1):
        cell = ws_calendar.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx > 1:
            cell.number_format = '"$"#,##0.00'

# Totals
ws_calendar['A10'] = "Monthly Total"
ws_calendar['A10'].font = Font(bold=True)
for col in range(2, 7):
    ws_calendar.cell(row=10, column=col, value=f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}8)')
    ws_calendar.cell(row=10, column=col).number_format = '"$"#,0.00'
    ws_calendar.cell(row=10, column=col).font = money_font

# Key Dates
ws_calendar['A12'] = "IMPORTANT DATES"
ws_calendar['A12'].font = subtitle_font

ws_calendar['A13'] = "Date"
ws_calendar['B13'] = "Bill/Expense"
ws_calendar['C13'] = "Amount"
ws_calendar['D13'] = "Status"
for col in range(1, 5):
    ws_calendar.cell(row=13, column=col).font = header_font
    ws_calendar.cell(row=13, column=col).fill = header_fill
    ws_calendar.cell(row=13, column=col).border = thin_border

important_dates = [
    ('June 1', 'Rent', 1500, 'Paid'),
    ('June 3', 'Electric Bill', 85, 'Paid'),
    ('June 10', 'Internet', 65, 'Paid'),
    ('June 15', 'Car Payment', 477, 'Due'),
    ('June 20', 'Student Loan', 357, 'Due'),
    ('June 25', 'Credit Card Due', 500, 'Due'),
]

for row_idx, date_info in enumerate(important_dates, 14):
    for col_idx, value in enumerate(date_info, 1):
        cell = ws_calendar.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 3:
            cell.number_format = '"$"#,##0.00'
        if col_idx == 4 and value == 'Paid':
            cell.fill = light_green_fill
        elif col_idx == 4 and value == 'Due':
            cell.fill = light_yellow_fill

# Daily Spending Tracker
ws_calendar['F3'] = "DAILY SPENDING TRACKER"
ws_calendar['F3'].font = subtitle_font

ws_calendar['F4'] = "Date"
ws_calendar['G4'] = "Spent"
ws_calendar['H4'] = "Running Total"
for col in range(6, 9):
    ws_calendar.cell(row=4, column=col).font = header_font
    ws_calendar.cell(row=4, column=col).fill = header_fill
    ws_calendar.cell(row=4, column=col).border = thin_border

for day in range(1, 31):
    row = day + 4
    ws_calendar.cell(row=row, column=6, value=f'June {day}')
    ws_calendar.cell(row=row, column=6).border = thin_border
    ws_calendar.cell(row=row, column=7).border = thin_border
    ws_calendar.cell(row=row, column=7).number_format = '"$"#,##0.00'
    if day > 1:
        ws_calendar.cell(row=row, column=8, value=f'=H{row-1}+G{row}')
    else:
        ws_calendar.cell(row=row, column=8, value=f'=G{row}')
    ws_calendar.cell(row=row, column=8).number_format = '"$"#,##0.00'
    ws_calendar.cell(row=row, column=8).border = thin_border

# Column widths
for col in range(1, 9):
    ws_calendar.column_dimensions[get_column_letter(col)].width = 15

# Set the Dashboard as the active sheet
wb.active = wb['Dashboard']

# Save the workbook
wb.save('Personal_Finance_Manager.xlsx')
print("Excel file created successfully: Personal_Finance_Manager.xlsx")
print("\nSheets included:")
print("1. Dashboard - Overview of your finances")
print("2. Income - Income sources and budget categories")
print("3. Expenses - Detailed expense tracking with categories")
print("4. Loans - Loan management and amortization")
print("5. Savings - Savings goals and progress tracking")
print("6. Charts - Visual representations of your finances")
print("7. Net Worth - Assets, liabilities, and net worth")
print("8. Monthly Budget - Weekly/daily budget calendar")
